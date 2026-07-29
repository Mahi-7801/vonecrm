import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Token Tracking"

# Define styles
header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

data_font = Font(name='Calibri', size=10)
data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
center_alignment = Alignment(horizontal='center', vertical='center')

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Column headers
headers = [
    'Token ID',
    'Token Name',
    'Token Type',
    'Created Date',
    'Expiration Date',
    'Assigned To',
    'Purpose',
    'Verification Status',
    'Verified By',
    'Verification Date',
    'Notes',
    'Last Updated'
]

# Column widths
col_widths = [12, 20, 15, 14, 14, 16, 25, 18, 16, 14, 30, 14]

# Write headers
for col, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col)].width = width

# Freeze header row
ws.freeze_panes = 'A2'

# Add data validation for Token Type
token_type_validation = DataValidation(
    type="list",
    formula1='"API Key,Access Token,Refresh Token,OAuth Token,JWT,Bearer Token,Session Token"',
    allow_blank=True
)
token_type_validation.error = "Please select a valid token type"
token_type_validation.errorTitle = "Invalid Token Type"
ws.add_data_validation(token_type_validation)
token_type_validation.add(f'C2:C1000')

# Add data validation for Verification Status
status_validation = DataValidation(
    type="list",
    formula1='"Pending,Approved,Rejected,Expired,Revoked"',
    allow_blank=True
)
status_validation.error = "Please select a valid verification status"
status_validation.errorTitle = "Invalid Status"
ws.add_data_validation(status_validation)
status_validation.add(f'H2:H1000')

# Add sample data rows
sample_data = [
    ['TKN-001', 'Production API Key', 'API Key', '2024-01-15', '2025-01-15', 'John Smith', 'Main production system access', 'Approved', 'Jane Doe', '2024-01-16', 'Standard production key', '2024-01-16'],
    ['TKN-002', 'Staging OAuth Token', 'OAuth Token', '2024-01-20', '2024-07-20', 'Mike Johnson', 'Staging environment testing', 'Approved', 'Jane Doe', '2024-01-21', 'For QA testing only', '2024-01-21'],
    ['TKN-003', 'Dev Access Token', 'Access Token', '2024-01-25', '2024-04-25', 'Sarah Wilson', 'Development environment', 'Pending', '', '', 'Awaiting security review', '2024-01-25'],
]

for row_num, row_data in enumerate(sample_data, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.font = data_font
        cell.alignment = data_alignment if col_num in [2, 7, 11] else center_alignment
        cell.border = thin_border

# Add second sheet for summary/dashboard
ws2 = wb.create_sheet("Summary")

# Summary headers
summary_headers = ['Metric', 'Count']
summary_data = [
    ['Total Tokens', f'=COUNTA(\'Token Tracking\'!A:A)-1'],
    ['Approved Tokens', f'=COUNTIF(\'Token Tracking\'!H:H,"Approved")'],
    ['Pending Tokens', f'=COUNTIF(\'Token Tracking\'!H:H,"Pending")'],
    ['Rejected Tokens', f'=COUNTIF(\'Token Tracking\'!H:H,"Rejected")'],
    ['Expired Tokens', f'=COUNTIF(\'Token Tracking\'!H:H,"Expired")'],
    ['Revoked Tokens', f'=COUNTIF(\'Token Tracking\'!H:H,"Revoked")'],
]

# Write summary headers
for col, header in enumerate(summary_headers, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 15

# Write summary data
for row_num, (metric, count) in enumerate(summary_data, 2):
    ws2.cell(row=row_num, column=1, value=metric).font = data_font
    ws2.cell(row=row_num, column=1).border = thin_border
    ws2.cell(row=row_num, column=2, value=count).font = data_font
    ws2.cell(row=row_num, column=2).alignment = center_alignment
    ws2.cell(row=row_num, column=2).border = thin_border

# Auto-fit summary column
ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 15

# Save workbook
output_path = 'Token_Tracking_Template.xlsx'
wb.save(output_path)
print(f"Template created successfully: {output_path}")